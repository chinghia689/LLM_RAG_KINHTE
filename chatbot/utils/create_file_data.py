import csv
from openpyxl import Workbook, load_workbook
import os


def save_to_csv(question: str, answer: str, docs: str, filename: str = "output.csv"):
    """
    Lưu dữ liệu vào file CSV (ghi đè nếu tồn tại).

    Args:
        question (str): Câu hỏi.
        answer (str): Câu trả lời.
        docs (str): Context/tài liệu liên quan.
        filename (str, optional): Tên file CSV. Mặc định "output.csv".
    """
    with open(filename, mode="w", newline="", encoding="utf-8") as file:
        writer = csv.writer(file)
        # Ghi header
        writer.writerow(["question", "answer", "docs"])
        # Ghi dữ liệu 1 dòng
        writer.writerow([question, answer, docs])


def save_to_excel(data_array: list, filename: str = "output.xlsx"):
    """
    Lưu danh sách dữ liệu vào file Excel (.xlsx).

    Args:
        data_array (list[dict]): Danh sách dict chứa các trường:
            - question
            - ground_truth
            - contexts_ground_truth
            - answer
            - contexts_answer
            - metadata
            - run_time
        filename (str, optional): Đường dẫn + tên file xuất ra.
                                  Mặc định "output.xlsx".
    """
    # Tạo folder nếu chưa tồn tại
    os.makedirs(os.path.dirname(filename), exist_ok=True)

    wb = Workbook()
    ws = wb.active

    # Ghi dòng tiêu đề
    ws.append(
        [
            "question",
            "ground_truth",
            "contexts_ground_truth",
            "answer",
            "contexts_answer",
            "metadata",
            "run_time",
        ]
    )

    # Ghi dữ liệu từng dòng
    for item in data_array:
        question = item.get("question", "")
        ground_truth = item.get("ground_truth", "")
        contexts_ground_truth = item.get("contexts_ground_truth", "")
        answer = item.get("answer", "")
        contexts_answer = item.get("contexts_answer", "")
        metadata = item.get("metadata", "")
        run_time = item.get("run_time", "")

        ws.append(
            [
                str(question),
                str(ground_truth),
                str(contexts_ground_truth),
                str(answer),
                str(contexts_answer),
                str(metadata),
                str(run_time),
            ]
        )

    wb.save(filename)


def read_excel(filename: str) -> list:
    """
    Đọc file Excel (.xlsx) thành list[dict].

    Args:
        filename (str): Đường dẫn file Excel.

    Returns:
        list[dict]: Danh sách các dòng, mỗi dòng là dict theo header.

    Raises:
        FileNotFoundError: Nếu file không tồn tại.
    """
    if not os.path.exists(filename):
        raise FileNotFoundError(f"Không tìm thấy file: {filename}")

    wb = load_workbook(filename)
    ws = wb.active

    data = []
    # Đọc dòng tiêu đề (header)
    headers = [cell.value for cell in ws[1]]

    # Đọc từng dòng dữ liệu, bắt đầu từ hàng 2
    for row in ws.iter_rows(min_row=2, values_only=True):
        data.append(dict(zip(headers, row)))

    return data
